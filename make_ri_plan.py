"""Generate RI planning XLSX for CTO review."""
import json, os, sys
from datetime import date, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# RDS pricing tables & storage-cost helper are shared with analyze.py — see
# rds_pricing.py so the two never drift apart.
from rds_pricing import (
    RDS_ONDEMAND_HOURLY as OD,
    RDS_RI_1YR_HOURLY as RI_1YR,
    RDS_RI_NU as NU,
    RDS_RAM_GB as RAM,
    rds_storage_monthly as storage_monthly,
)

# ── Snapshot path ─────────────────────────────────────────────────────────────
def _find_latest():
    base = Path("data")
    if not base.exists():
        sys.exit("No ./data directory found. Run pull first.")
    for date_dir in sorted(base.iterdir(), key=lambda d: d.name, reverse=True):
        if not date_dir.is_dir():
            continue
        if (date_dir / "infra/rds/instances.json").exists():
            return str(date_dir)
        for profile_dir in sorted(date_dir.iterdir(), key=lambda d: d.name):
            if (profile_dir / "infra/rds/instances.json").exists():
                return str(profile_dir)
    sys.exit("No snapshots with RDS data found in ./data/")

DATA = sys.argv[1] if len(sys.argv) > 1 else _find_latest()
SNAP_LABEL = "/".join(Path(DATA).parts[-2:]) if len(Path(DATA).parts) >= 2 else Path(DATA).name
TODAY = date.today()

with open(f"{DATA}/infra/rds/instances.json") as f:
    _inst = json.load(f)
with open(f"{DATA}/infra/rds/reserved_instances.json") as f:
    _ri = json.load(f)

if "DBInstances" not in _inst:
    sys.exit(f"No RDS instance data in {DATA} — pull may have failed for this profile.")

instances_raw = _inst["DBInstances"]
ri_raw        = _ri.get("ReservedDBInstances", [])

# Active RIs only
active_ri = [r for r in ri_raw if r["State"] == "active"]

# ── RI pool: compute NUs by engine family ─────────────────────────────────────
# Key: (family e.g. "m7i", engine_group e.g. "mysql") → available NUs
ri_pool = {}
ri_detail = []
for r in active_ri:
    cls = r["DBInstanceClass"]
    count = r["DBInstanceCount"]
    nu = NU.get(cls, 0) * count
    engine = r.get("ProductDescription", "").lower()
    engine_grp = "postgresql" if "postgres" in engine else "mysql"
    family = cls.split(".")[1]  # "m7i", "t3", etc.
    key = (family, engine_grp)
    ri_pool[key] = ri_pool.get(key, 0) + nu

    start = r.get("StartTime", "")[:10]
    duration_yrs = r.get("Duration", 31536000) // 31536000
    if start:
        start_dt = datetime.strptime(start, "%Y-%m-%d")
        from dateutil.relativedelta import relativedelta
        expiry = (start_dt + relativedelta(years=duration_yrs)).strftime("%Y-%m-%d")
    else:
        expiry = "unknown"
    ri_detail.append({
        "id": r["ReservedDBInstanceId"],
        "class": cls,
        "count": count,
        "nu_each": NU.get(cls, 0),
        "nu_total": nu,
        "engine": engine_grp,
        "family": family,
        "payment": r.get("OfferingType", ""),
        "start": start,
        "expiry": expiry,
        "state": r["State"],
    })

# ── Assign RI coverage to instances ──────────────────────────────────────────
# Sort instances: exact family+engine matches first, largest first
def engine_grp(e):
    return "postgresql" if "postgres" in e.lower() else "mysql"

instances = []
for db in instances_raw:
    cls = db["DBInstanceClass"]
    eng = engine_grp(db.get("Engine", ""))
    family = cls.split(".")[1]
    nu_needed = NU.get(cls, 0)
    storage = db.get("AllocatedStorage", 0)
    storage_type = db.get("StorageType", "gp2")
    iops = db.get("Iops") or 0
    throughput = db.get("StorageThroughput") or 0
    storage_mo = storage_monthly(storage, storage_type, iops, throughput)
    storage_daily = storage_mo / 30
    od_hr = OD.get(cls, 0)
    ri_hr = RI_1YR.get(cls)

    # pull from pool
    key = (family, eng)
    pool = ri_pool.get(key, 0)
    nu_covered = min(pool, nu_needed)
    ri_pool[key] = max(0, pool - nu_needed)
    frac = nu_covered / nu_needed if nu_needed else 0

    if ri_hr and frac > 0:
        eff_hr = frac * ri_hr + (1 - frac) * od_hr
    else:
        eff_hr = od_hr

    # Use 730 hours/month (AWS billing standard)
    monthly_est     = eff_hr * 730 + storage_mo
    monthly_od      = od_hr  * 730 + storage_mo
    monthly_full_ri = (ri_hr * 730 + storage_mo) if ri_hr else None

    instances.append({
        "id": db["DBInstanceIdentifier"],
        "class": cls,
        "engine": db.get("Engine", ""),
        "family": family,
        "engine_grp": eng,
        "ram_gb": RAM.get(cls, "?"),
        "storage_gb": storage,
        "multi_az": db.get("MultiAZ", False),
        "status": db.get("DBInstanceStatus", ""),
        "nu": nu_needed,
        "nu_covered": nu_covered,
        "ri_pct": round(frac * 100),
        "daily_est": eff_hr * 24 + storage_daily,  # kept for sort only
        "monthly_est": monthly_est,
        "monthly_od": monthly_od,
        "monthly_full_ri": monthly_full_ri,
        "monthly_saving_if_full_ri": monthly_od - (monthly_full_ri or monthly_od),
        "ri_hr": ri_hr,
        "od_hr": od_hr,
    })

# Sort by daily cost desc
instances.sort(key=lambda x: x["daily_est"], reverse=True)

# ── Recommendations ───────────────────────────────────────────────────────────
recommendations = []

# 1. Instances with partial or no RI coverage — one entry per instance
for inst in instances:
    if inst["ri_pct"] < 100 and inst["ri_hr"] and inst["monthly_saving_if_full_ri"] > 100:
        gap_nu = inst["nu"] - inst["nu_covered"]
        saving = inst["monthly_saving_if_full_ri"]
        ri_status = f"{inst['ri_pct']}% covered" if inst["ri_pct"] > 0 else "No RI"
        action = (f"Buy RI to cover remaining {gap_nu} NUs ({inst['engine_grp']})"
                  if inst["ri_pct"] > 0 else
                  f"Buy 1× {inst['class']} RI ({inst['engine_grp']})")
        recommendations.append({
            "priority": "High" if saving > 500 else "Medium",
            "instance": inst["id"],
            "current_class": inst["class"],
            "current_ri": ri_status,
            "action": action,
            "monthly_saving": saving,
            "note": "",
        })

# 2. Expiry warnings (within 6 months)
for ri in ri_detail:
    if ri["expiry"] == "unknown":
        continue
    exp_dt = datetime.strptime(ri["expiry"], "%Y-%m-%d").date()
    days_left = (exp_dt - TODAY).days
    if days_left < 180:
        recommendations.append({
            "priority": "High" if days_left < 90 else "Medium",
            "instance": "—",
            "current_class": ri["class"],
            "current_ri": f"Expires {ri['expiry']} ({days_left}d left)",
            "action": f"Renew RI — covers {ri['nu_total']} NUs ({ri['engine']})",
            "monthly_saving": 0,
            "note": "Letting this expire creates a coverage gap from that date",
        })

recs_clean = sorted(recommendations, key=lambda x: (-x["monthly_saving"], x["priority"]))

# ── Styles ────────────────────────────────────────────────────────────────────
DARK   = "1A3C5E"
MID    = "1A5FA8"
LIGHT  = "D6E4F0"
CHALK  = "F0F4F8"
GREEN  = "177840"
RED    = "B52B2B"
ORANGE = "B05810"
WHITE  = "FFFFFF"
YELLOW = "FFF3CD"

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(bold=False, color=WHITE, size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border():
    s = Side(style="thin", color="C6D0DD")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, cols, widths=None):
    for i, (label, width) in enumerate(zip(cols, widths or [18]*len(cols)), 1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill = fill(DARK)
        c.font = font(bold=True, color=WHITE, size=9)
        c.alignment = align("center")
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(i)].width = width

def data_cell(ws, row, col, value, bold=False, color=WHITE, bg=None,
              h="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, size=9, name="Calibri")
    c.alignment = align(h, wrap=True)
    c.border = thin_border()
    if bg:
        c.fill = fill(bg)
    if num_fmt:
        c.number_format = num_fmt
    return c

def section_title(ws, row, title, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = fill(MID)
    c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.alignment = align("left")
    ws.row_dimensions[row].height = 20

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Current Instances ────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "RDS Instances"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A3"

section_title(ws1, 1, f"RDS Instances — Current State  (snapshot: {SNAP_LABEL})", 14)
cols1 = ["Instance ID", "Class", "Engine", "RAM (GB)", "Storage (GB)",
         "Multi-AZ", "NUs", "RI Coverage",
         "$/month (with RI)", "$/month (no RI)", "Saving /month", "% Saved", "Status"]
widths1 = [32, 18, 10, 10, 12, 10, 8, 14, 18, 18, 16, 10, 12]
header_row(ws1, 2, cols1, widths1)

for r_idx, inst in enumerate(instances, 3):
    ri_pct = inst["ri_pct"]
    bg = CHALK if r_idx % 2 == 0 else WHITE

    ri_color = "111111"
    if ri_pct == 100:   ri_color = GREEN
    elif ri_pct >= 50:  ri_color = ORANGE
    elif ri_pct == 0:   ri_color = RED

    monthly_with_ri = round(inst["monthly_est"], 2)
    monthly_no_ri   = round(inst["monthly_od"], 2)
    saving          = round(monthly_no_ri - monthly_with_ri, 2)
    pct_saved       = round((saving / monthly_no_ri * 100) if monthly_no_ri else 0, 1)

    data_cell(ws1, r_idx, 1,  inst["id"], bg=bg, color="111111")
    data_cell(ws1, r_idx, 2,  inst["class"], bg=bg, color="111111")
    data_cell(ws1, r_idx, 3,  inst["engine"], bg=bg, color="111111")
    data_cell(ws1, r_idx, 4,  inst["ram_gb"], bg=bg, color="111111", h="right")
    data_cell(ws1, r_idx, 5,  inst["storage_gb"], bg=bg, color="111111", h="right")
    data_cell(ws1, r_idx, 6,  "Yes" if inst["multi_az"] else "No", bg=bg, color="111111", h="center")
    data_cell(ws1, r_idx, 7,  inst["nu"] or "—", bg=bg, color="111111", h="right")
    ri_label = f"{ri_pct}% RI" if ri_pct > 0 else "On-demand"
    data_cell(ws1, r_idx, 8,  ri_label, bold=(ri_pct==100), bg=bg, color=ri_color, h="center")
    data_cell(ws1, r_idx, 9,  monthly_with_ri, bg=bg, color="111111",
              h="right", num_fmt='0.00')
    data_cell(ws1, r_idx, 10, monthly_no_ri, bg=bg, color="111111",
              h="right", num_fmt='0.00')
    # Saving cell — green if saving > 0, grey if zero
    saving_color = GREEN if saving > 0 else "888888"
    data_cell(ws1, r_idx, 11, saving if saving > 0 else "—",
              bold=(saving > 100), bg=bg, color=saving_color,
              h="right", num_fmt=('0.00' if saving > 0 else None))
    data_cell(ws1, r_idx, 12, f"{pct_saved}%" if saving > 0 else "—",
              bold=(saving > 100), bg=bg, color=saving_color, h="center")
    data_cell(ws1, r_idx, 13, inst["status"], bg=bg, color="111111", h="center")

# Total row
total_saving   = round(sum(i["monthly_od"] - i["monthly_est"] for i in instances), 2)
tr = len(instances) + 3
for col in range(1, 14):
    c = ws1.cell(row=tr, column=col)
    c.fill = fill(LIGHT)
    c.border = thin_border()
    c.font = Font(bold=True, color="111111", size=9, name="Calibri")
ws1.cell(row=tr, column=1, value="TOTAL").alignment = align("left")
c9  = ws1.cell(row=tr, column=9,  value=round(sum(i["monthly_est"] for i in instances), 2))
c10 = ws1.cell(row=tr, column=10, value=round(sum(i["monthly_od"]  for i in instances), 2))
c11 = ws1.cell(row=tr, column=11, value=total_saving)
for c, fmt in [(c9, '0.00'), (c10, '0.00'), (c11, '0.00')]:
    c.number_format = fmt; c.alignment = align("right")
c11.font = Font(bold=True, color=GREEN, size=9, name="Calibri")

# Note row
nr = tr + 1
ws1.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=13)
nc = ws1.cell(row=nr, column=1,
    value="Note: $/month = hourly rate × 730 hrs (AWS billing standard) + storage. "
          "RI rates: 1-yr No Upfront, ap-southeast-1. "
          "Storage: gp2 = $0.138/GB; gp3 = $0.138/GB + $0.024/IOPS-mo (above 3,000 free) + $0.096/MiBps-mo (above 125 free). "
          "All rates confirmed via AWS pricing API.")
nc.font = Font(italic=True, color="5E7490", size=8, name="Calibri")
nc.alignment = align("left", wrap=True)
ws1.row_dimensions[nr].height = 28

# ── Sheet 2: Reserved Instances ───────────────────────────────────────────────
ws2 = wb.create_sheet("Reserved Instances")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"

section_title(ws2, 1, f"Active Reserved Instances  (snapshot: {SNAP_LABEL})", 9)
cols2 = ["RI ID", "Class", "Engine", "Count", "NUs (each)", "NUs (total)",
         "Payment", "Start Date", "Expiry Date"]
widths2 = [38, 18, 12, 8, 12, 12, 14, 14, 14]
header_row(ws2, 2, cols2, widths2)

for r_idx, ri in enumerate(ri_detail, 3):
    bg = CHALK if r_idx % 2 == 0 else WHITE
    # Color expiry
    try:
        exp_dt = datetime.strptime(ri["expiry"], "%Y-%m-%d").date()
        days_left = (exp_dt - TODAY).days
        exp_color = RED if days_left < 90 else (ORANGE if days_left < 180 else GREEN)
    except:
        days_left = 999; exp_color = "111111"

    data_cell(ws2, r_idx, 1, ri["id"], bg=bg, color="111111")
    data_cell(ws2, r_idx, 2, ri["class"], bg=bg, color="111111")
    data_cell(ws2, r_idx, 3, ri["engine"], bg=bg, color="111111")
    data_cell(ws2, r_idx, 4, ri["count"], bg=bg, color="111111", h="center")
    data_cell(ws2, r_idx, 5, ri["nu_each"], bg=bg, color="111111", h="right")
    data_cell(ws2, r_idx, 6, ri["nu_total"], bg=bg, color="111111", h="right")
    data_cell(ws2, r_idx, 7, ri["payment"], bg=bg, color="111111", h="center")
    data_cell(ws2, r_idx, 8, ri["start"], bg=bg, color="111111", h="center")
    exp_label = f"{ri['expiry']} ({days_left}d left)" if days_left < 999 else ri["expiry"]
    data_cell(ws2, r_idx, 9, exp_label, bold=(days_left<90), bg=bg,
              color=exp_color, h="center")

# NU summary note
sr = len(ri_detail) + 3
ws2.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=9)
total_nu_mysql = sum(r["nu_total"] for r in ri_detail if r["engine"]=="mysql")
total_nu_pg = sum(r["nu_total"] for r in ri_detail if r["engine"]=="postgresql")
sc = ws2.cell(row=sr, column=1,
    value=f"Total active NUs available:  MySQL = {total_nu_mysql} NUs  |  PostgreSQL = {total_nu_pg} NUs")
sc.font = Font(bold=True, color=DARK, size=9, name="Calibri")
sc.alignment = align("left")
sc.fill = fill(LIGHT)
sc.border = thin_border()

# ── Sheet 3: RI Recommendations ───────────────────────────────────────────────
ws3 = wb.create_sheet("RI Recommendations")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A3"

section_title(ws3, 1, f"RI Planning Recommendations  (as of {TODAY})", 7)
cols3 = ["Priority", "Instance", "Current Class", "Current RI Status",
         "Recommended Action", "Est. Monthly Saving", "Notes"]
widths3 = [10, 32, 18, 20, 42, 20, 36]
header_row(ws3, 2, cols3, widths3)

for r_idx, rec in enumerate(recs_clean, 3):
    bg = CHALK if r_idx % 2 == 0 else WHITE
    pri_color = RED if rec["priority"] == "High" else ORANGE

    data_cell(ws3, r_idx, 1, rec["priority"], bold=True, bg=bg,
              color=pri_color, h="center")
    data_cell(ws3, r_idx, 2, rec["instance"], bg=bg, color="111111")
    data_cell(ws3, r_idx, 3, rec["current_class"], bg=bg, color="111111")
    data_cell(ws3, r_idx, 4, rec["current_ri"], bg=bg, color="111111")
    data_cell(ws3, r_idx, 5, rec["action"], bg=bg, color="111111")
    saving = rec["monthly_saving"]
    saving_cell = data_cell(ws3, r_idx, 6,
        round(saving, 0) if saving > 0 else "—",
        bold=(saving > 500), bg=bg,
        color=GREEN if saving > 0 else "111111", h="right")
    if saving > 0:
        saving_cell.number_format = '0'
    data_cell(ws3, r_idx, 7, rec["note"], bg=bg, color="5E7490")
    ws3.row_dimensions[r_idx].height = 30

# Footer note
fn = len(recs_clean) + 3
ws3.merge_cells(start_row=fn, start_column=1, end_row=fn, end_column=7)
fc = ws3.cell(row=fn, column=1,
    value="Monthly saving estimates based on 1-yr No Upfront RI vs on-demand rates, ap-southeast-1. "
          "Verify prices at aws.amazon.com/rds/pricing before purchase.")
fc.font = Font(italic=True, color="5E7490", size=8, name="Calibri")
fc.alignment = align("left", wrap=True)
ws3.row_dimensions[fn].height = 28

# ── Save ──────────────────────────────────────────────────────────────────────
out = f"ri_plan_{SNAP_LABEL.replace('/', '_')}.xlsx"
wb.save(out)
print(f"Saved: {out}")
